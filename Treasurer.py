import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Create a new workbook and remove the default sheet if needed
wb = Workbook()

# ---------------------
# Setup Sheet
# ---------------------
ws_setup = wb.active
ws_setup.title = "Setup"

# Add mapping table headers and a few example mappings
ws_setup["A1"] = "Bank Export Field"
ws_setup["B1"] = "Target Field"
ws_setup["A2"] = "Trans Date"
ws_setup["B2"] = "Date"
ws_setup["A3"] = "Description"
ws_setup["B3"] = "Description"
ws_setup["A4"] = "Amount"
ws_setup["B4"] = "Amount"

# Add a list of categories starting in column D
ws_setup["D1"] = "Categories"
categories = [
    "Treasurer Funds",
    "Charity",
    "Benevolent Funds",
    "LOI",
    "Dining Fees",
    "Tyler Payment",
    "Gavel Collection",
    "Raffle Collection"
]
for i, cat in enumerate(categories, start=2):
    ws_setup[f"D{i}"] = cat

# ---------------------
# RawData Sheet
# ---------------------
ws_raw = wb.create_sheet("RawData")
# Headers for raw bank export data
ws_raw["A1"] = "Trans Date"
ws_raw["B1"] = "Description"
ws_raw["C1"] = "Amount"
# Example row (this sheet is meant for pasting/importing bank data)
ws_raw.append(["2025-01-15", "Sample Bank Transaction", 100.00])

# Convert RawData range into a Table for dynamic expansion (optional)
raw_table = Table(displayName="RawDataTable", ref=f"A1:C{ws_raw.max_row}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
raw_table.tableStyleInfo = style
ws_raw.add_table(raw_table)

# ---------------------
# Transactions Sheet
# ---------------------
ws_trans = wb.create_sheet("Transactions")
# Define headers for the unified transaction log
ws_trans.append(["Date", "Description", "Amount", "Category", "Payment Method"])

# Instructions:
# - For bank transactions: you can use formulas or Power Query to reference the RawData table.
# - For cash/cheque transactions: enter them manually in the rows below.
#
# For example, you might set up a formula in the Date column (cell A2) like:
#   =IF(ROW()-1<=ROWS(RawDataTable[#All]), INDEX(RawDataTable[Trans Date], ROW()-1), "")
# Adjust formulas as needed for your reporting setup.

# ---------------------
# Members Sheet
# ---------------------
ws_members = wb.create_sheet("Members")
ws_members.append(["Member ID", "Name", "Sub Account"])
# Example member
ws_members.append([1, "John Doe", "Treasurer Funds"])

# ---------------------
# Summary Sheet
# ---------------------
ws_summary = wb.create_sheet("Summary")
ws_summary.append(["Category", "Total"])
# Example formulas (these assume you have bank data and manual entries in Transactions)
ws_summary.append(["Treasurer Funds", '=SUMIF(Transactions!D:D, "Treasurer Funds", Transactions!C:C)'])
ws_summary.append(["Charity", '=SUMIF(Transactions!D:D, "Charity", Transactions!C:C)'])
ws_summary.append(["Benevolent Funds", '=SUMIF(Transactions!D:D, "Benevolent Funds", Transactions!C:C)'])
ws_summary.append(["LOI", '=SUMIF(Transactions!D:D, "LOI", Transactions!C:C)'])
ws_summary.append(["Dining Fees", '=SUMIF(Transactions!D:D, "Dining Fees", Transactions!C:C)'])
ws_summary.append(["Tyler Payment", '=SUMIF(Transactions!D:D, "Tyler Payment", Transactions!C:C)'])
ws_summary.append(["Gavel Collection", '=SUMIF(Transactions!D:D, "Gavel Collection", Transactions!C:C)'])
ws_summary.append(["Raffle Collection", '=SUMIF(Transactions!D:D, "Raffle Collection", Transactions!C:C)'])

# Save the workbook
output_filename = "TreasurerAccounts_Template.xlsx"
wb.save(output_filename)
print(f"Template saved as {output_filename}")
